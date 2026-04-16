"""Rotas REST para a base de conhecimento (Knowledge Base).

CRUD de documentos: upload com parsing multi-formato, listagem,
download do arquivo original e exclusão.

Todas as rotas exigem sessão Better Auth válida via get_current_user.
"""

import csv
import io
from pathlib import Path

import structlog
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from gyros_os.rag.ingest import ingest_text
from gyros_os.server.dependencies import get_current_user
from gyros_os.shared.db import get_pool
from gyros_os.shared.org import get_default_org_id

logger = structlog.get_logger()

router = APIRouter(prefix="/api/kb", tags=["knowledge-base"])

ALLOWED_EXTENSIONS = {".txt", ".md", ".csv", ".pdf", ".docx", ".xlsx"}


async def _parse_file(content: bytes, filename: str) -> str:
    """Extrai texto de um arquivo com base na extensão."""
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            415,
            f"Tipo não suportado: {ext}. "
            f"Aceitos: {', '.join(sorted(ALLOWED_EXTENSIONS))}",
        )

    if ext in (".txt", ".md"):
        return content.decode("utf-8", errors="replace")

    if ext == ".csv":
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return "\n".join(", ".join(row) for row in reader)

    if ext == ".pdf":
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(content))
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)

    if ext == ".docx":
        from docx import Document

        doc = Document(io.BytesIO(content))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())

    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"=== Planilha: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_text.strip():
                    parts.append(row_text)
        return "\n".join(parts)

    raise HTTPException(415, f"Parser não implementado para {ext}")


@router.post("/upload")
async def upload_document(
    file: UploadFile,
    user: dict = Depends(get_current_user),
) -> dict:
    """Faz upload de documento, parseia texto e indexa na KB.

    Aceita .txt, .md, .csv, .pdf, .docx e .xlsx. O arquivo original é
    armazenado em kb_docs.file_data para download posterior.
    """
    filename = file.filename or "unnamed"
    content_bytes = await file.read()

    if not content_bytes:
        raise HTTPException(400, "Arquivo vazio")

    text_content = await _parse_file(content_bytes, filename)
    if not text_content.strip():
        raise HTTPException(422, "Nenhum texto extraído do arquivo")

    pool = await get_pool()
    org_id = await get_default_org_id(pool)

    result = await ingest_text(
        org_id=org_id,
        source_type="upload",
        source_ref=filename,
        title=filename,
        content=text_content,
    )

    doc_id = result["doc_id"]

    # Salva arquivo original para download
    async with pool.connection() as conn:
        await conn.execute(
            """
            UPDATE kb_docs
            SET file_data = %s, file_name = %s, file_size = %s, mime_type = %s
            WHERE id = %s
            """,
            (content_bytes, filename, len(content_bytes), file.content_type, doc_id),
        )
        await conn.commit()

    logger.info(
        "kb_upload_complete",
        doc_id=doc_id,
        filename=filename,
        num_chunks=result["num_chunks"],
        user_id=user["user_id"],
    )

    return {
        "doc_id": doc_id,
        "filename": filename,
        "num_chunks": result["num_chunks"],
        "status": "ready",
    }


@router.get("/docs")
async def list_documents(
    user: dict = Depends(get_current_user),
) -> list[dict]:
    """Lista documentos da organização (sem file_data)."""
    pool = await get_pool()
    org_id = await get_default_org_id(pool)

    async with pool.connection() as conn:
        cursor = await conn.execute(
            """
            SELECT id, title, source_type, metadata, created_at,
                   file_name, file_size, mime_type,
                   (SELECT count(*) FROM kb_chunks
                    WHERE doc_id = kb_docs.id) AS chunk_count
            FROM kb_docs
            WHERE organization_id = %s
            ORDER BY created_at DESC
            """,
            (str(org_id),),
        )
        rows = await cursor.fetchall()

    return [
        {
            "id": str(r[0]),
            "title": r[1],
            "source_type": r[2],
            "metadata": r[3],
            "created_at": r[4].isoformat() if r[4] else None,
            "file_name": r[5],
            "file_size": r[6],
            "mime_type": r[7],
            "chunk_count": r[8],
        }
        for r in rows
    ]


@router.get("/docs/{doc_id}/download")
async def download_document(
    doc_id: str,
    user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Retorna o arquivo original para download."""
    pool = await get_pool()
    org_id = await get_default_org_id(pool)

    async with pool.connection() as conn:
        cursor = await conn.execute(
            """
            SELECT file_data, file_name, mime_type
            FROM kb_docs
            WHERE id = %s AND organization_id = %s
            """,
            (doc_id, str(org_id)),
        )
        row = await cursor.fetchone()

    if not row:
        raise HTTPException(404, "Document not found")

    file_data, file_name, mime_type = row[0], row[1], row[2]
    if not file_data:
        raise HTTPException(404, "Arquivo original não disponível")

    return StreamingResponse(
        io.BytesIO(file_data),
        media_type=mime_type or "application/octet-stream",
        headers={
            "Content-Disposition": f'attachment; filename="{file_name or "download"}"',
            "Content-Length": str(len(file_data)),
        },
    )


@router.delete("/docs/{doc_id}")
async def delete_document(
    doc_id: str,
    user: dict = Depends(get_current_user),
) -> dict:
    """Remove documento e seus chunks."""
    pool = await get_pool()
    org_id = await get_default_org_id(pool)

    async with pool.connection() as conn:
        await conn.execute(
            "DELETE FROM kb_chunks WHERE doc_id = %s",
            (doc_id,),
        )
        cursor = await conn.execute(
            "DELETE FROM kb_docs WHERE id = %s AND organization_id = %s RETURNING id",
            (doc_id, str(org_id)),
        )
        row = await cursor.fetchone()
        await conn.commit()

    if not row:
        raise HTTPException(404, "Document not found")

    logger.info("kb_doc_deleted", doc_id=doc_id, user_id=user["user_id"])
    return {"deleted": True, "doc_id": doc_id}
